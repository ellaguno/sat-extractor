"""Generación de reportes Excel con openpyxl."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..db.repository import Repository
from ..fiscal import calcular_impuestos_mensuales
from ..fiscal.clasificador import ClasificadorDeducciones
from ..models import TIPO_COMPROBANTE

# Estilos
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12, color="2F5496")
TOTAL_FONT = Font(bold=True, size=11)
CURRENCY_FMT = '#,##0.00'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

CFDI_COLUMNS = [
    ("Fecha", 13),
    ("UUID", 38),
    ("RFC Emisor", 15),
    ("Nombre Emisor", 30),
    ("RFC Receptor", 15),
    ("Nombre Receptor", 30),
    ("Concepto", 40),
    ("Tipo", 10),
    ("SubTotal", 14),
    ("Descuento", 12),
    ("Total", 14),
    ("IVA Trasladado", 14),
    ("ISR Retenido", 13),
    ("IVA Retenido", 13),
    ("Moneda", 8),
    ("Método Pago", 12),
    ("Forma Pago", 11),
    ("Estado", 10),
]

MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


# Estilos para análisis fiscal
DEDUCIBLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NO_DEDUCIBLE_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
ALERTA_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
FISCAL_GREEN = Font(color="2E7D32")
FISCAL_RED = Font(color="C62828")
FISCAL_ORANGE = Font(color="E65100")
SUGERENCIA_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
SUGERENCIA_FONT = Font(bold=True, size=12, color="283593")


class ExcelExporter:
    def __init__(self, db: Repository, regimen: str = "612"):
        self.db = db
        self.regimen = regimen

    def annual_report(self, year: int, output_dir: Path) -> Path:
        """Genera reporte anual: pestaña resumen + una pestaña por mes."""
        output_dir.mkdir(parents=True, exist_ok=True)
        filename = f"reporte_anual_{year:04d}.xlsx"
        output_path = output_dir / filename

        wb = Workbook()

        # Primera pestaña: Resumen Anual
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Anual"
        self._write_annual_summary(ws_resumen, year)

        # Una pestaña por cada mes
        for month in range(1, 13):
            ws = wb.create_sheet(MESES[month])
            self._write_month_sheet(ws, year, month)

        # Pestaña de Análisis Fiscal
        fecha_inicio_anual = date(year, 1, 1)
        fecha_fin_anual = date(year + 1, 1, 1)
        ws_fiscal = wb.create_sheet("Análisis Fiscal")
        self._write_fiscal_analysis(ws_fiscal, fecha_inicio_anual, fecha_fin_anual, str(year))

        # Pestaña de Sugerencias
        ws_sug = wb.create_sheet("Sugerencias")
        self._write_suggestions(ws_sug, fecha_inicio_anual, fecha_fin_anual, str(year))

        wb.save(str(output_path))
        return output_path

    def monthly_report(self, year: int, month: int, output_dir: Path) -> Path:
        """Genera reporte de un solo mes."""
        output_dir.mkdir(parents=True, exist_ok=True)
        filename = f"reporte_mensual_{year:04d}_{month:02d}.xlsx"
        output_path = output_dir / filename

        wb = Workbook()

        ws = wb.active
        titulo_mes = f"{MESES[month]} {year}"
        ws.title = titulo_mes
        self._write_month_sheet(ws, year, month)

        # Análisis fiscal del mes
        fecha_inicio = date(year, month, 1)
        fecha_fin = _next_month(year, month)
        ws_fiscal = wb.create_sheet("Análisis Fiscal")
        self._write_fiscal_analysis(ws_fiscal, fecha_inicio, fecha_fin, titulo_mes)

        # Sugerencias del mes
        ws_sug = wb.create_sheet("Sugerencias")
        self._write_suggestions(ws_sug, fecha_inicio, fecha_fin, titulo_mes)

        wb.save(str(output_path))
        return output_path

    def _write_month_sheet(self, ws, year: int, month: int):
        """Escribe una hoja con emitidas arriba y recibidas abajo."""
        fecha_inicio = date(year, month, 1)
        fecha_fin = _next_month(year, month)

        emitidas = self.db.search(
            tipo="emitida",
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            limit=10000,
        )
        recibidas = self.db.search(
            tipo="recibida",
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            limit=10000,
        )

        # Configurar anchos de columna
        for col_idx, (_, width) in enumerate(CFDI_COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row = 1

        # === SECCIÓN EMITIDAS ===
        row = self._write_section(ws, row, f"EMITIDAS - {MESES[month]} {year}", emitidas)

        # Espacio
        row += 2

        # === SECCIÓN RECIBIDAS ===
        row = self._write_section(ws, row, f"RECIBIDAS - {MESES[month]} {year}", recibidas)

        ws.freeze_panes = "A1"

    def _write_section(self, ws, start_row: int, title: str, cfdis) -> int:
        """Escribe una sección (título + headers + datos + subtotal).
        Retorna la siguiente fila disponible.
        """
        row = start_row
        num_cols = len(CFDI_COLUMNS)

        # Título de sección
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = SECTION_FONT
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = SECTION_FILL
        row += 1

        # Headers
        for col_idx, (name, _) in enumerate(CFDI_COLUMNS, 1):
            cell = ws.cell(row=row, column=col_idx, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Datos
        total_subtotal = 0.0
        total_descuento = 0.0
        total_total = 0.0
        total_iva_trasl = 0.0
        total_isr_ret = 0.0
        total_iva_ret = 0.0

        if not cfdis:
            ws.cell(row=row, column=1, value="(Sin registros)")
            ws.cell(row=row, column=1).font = Font(italic=True, color="888888")
            row += 1
        else:
            for c in cfdis:
                # Concepto: unir descripciones de todos los conceptos
                concepto = ""
                if c.conceptos:
                    descripciones = [con.descripcion for con in c.conceptos if con.descripcion]
                    concepto = " | ".join(descripciones)

                ws.cell(row=row, column=1, value=c.fecha.strftime("%d/%m/%Y"))
                ws.cell(row=row, column=2, value=c.uuid)
                ws.cell(row=row, column=3, value=c.rfc_emisor)
                ws.cell(row=row, column=4, value=c.nombre_emisor)
                ws.cell(row=row, column=5, value=c.rfc_receptor)
                ws.cell(row=row, column=6, value=c.nombre_receptor)
                ws.cell(row=row, column=7, value=concepto)
                ws.cell(row=row, column=8, value=TIPO_COMPROBANTE.get(c.tipo_comprobante, c.tipo_comprobante))

                subtotal = float(c.subtotal) if c.subtotal else 0
                descuento = float(c.descuento) if c.descuento else 0
                total = float(c.total) if c.total else 0
                iva_t = float(c.iva_trasladado) if c.iva_trasladado else 0
                isr_r = float(c.isr_retenido) if c.isr_retenido else 0
                iva_r = float(c.iva_retenido) if c.iva_retenido else 0

                for col, val in [
                    (9, subtotal), (10, descuento), (11, total),
                    (12, iva_t), (13, isr_r), (14, iva_r),
                ]:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.number_format = CURRENCY_FMT

                ws.cell(row=row, column=15, value=c.moneda)
                ws.cell(row=row, column=16, value=c.metodo_pago or "")
                ws.cell(row=row, column=17, value=c.forma_pago or "")
                ws.cell(row=row, column=18, value=c.estado)

                cancelado = c.estado != "Vigente"
                if cancelado:
                    # Marcar fila completa con texto tachado y gris
                    cancel_font = Font(strikethrough=True, color="999999")
                    for col in range(1, 19):
                        ws.cell(row=row, column=col).font = cancel_font
                else:
                    total_subtotal += subtotal
                    total_descuento += descuento
                    total_total += total
                    total_iva_trasl += iva_t
                    total_isr_ret += isr_r
                    total_iva_ret += iva_r

                row += 1

            # Fila de totales
            vigentes = sum(1 for c in cfdis if c.estado == "Vigente")
            cancelados = len(cfdis) - vigentes
            label = f"TOTAL ({vigentes} CFDIs"
            if cancelados:
                label += f", {cancelados} cancelado{'s' if cancelados > 1 else ''}"
            label += ")"
            ws.cell(row=row, column=7, value=label).font = TOTAL_FONT
            for col, val in [
                (9, total_subtotal), (10, total_descuento), (11, total_total),
                (12, total_iva_trasl), (13, total_isr_ret), (14, total_iva_ret),
            ]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.number_format = CURRENCY_FMT
                cell.font = TOTAL_FONT
            # Línea superior para totales
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = Border(
                    top=Side(style="thin", color="2F5496")
                )
            row += 1

        return row

    def _write_annual_summary(self, ws, year: int):
        """Escribe la pestaña de resumen anual."""
        # Título
        title_cell = ws.cell(row=1, column=1, value=f"Resumen Anual {year}")
        title_cell.font = Font(bold=True, size=16, color="2F5496")

        # === SECCIÓN EMITIDAS ===
        row = 3
        cell = ws.cell(row=row, column=1, value="EMITIDAS")
        cell.font = SECTION_FONT
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = SECTION_FILL
        row += 1

        emi_headers = ["Mes", "CFDIs", "Ingresos", "Egresos", "IVA Trasladado", "ISR Retenido", "IVA Retenido", "Total"]
        emi_widths = [14, 10, 16, 16, 16, 14, 14, 16]
        for col_idx, (h, w) in enumerate(zip(emi_headers, emi_widths), 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        row += 1

        emi_totals = [0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        for month in range(1, 13):
            se = self.db.monthly_summary(year, month, "emitida")
            ws.cell(row=row, column=1, value=MESES[month])
            values = [
                se["num_cfdis"], se["ingresos"], se["egresos"],
                se["iva_trasladado"], se["isr_retenido"], se["iva_retenido"],
                se["total"],
            ]
            for col_idx, val in enumerate(values):
                cell = ws.cell(row=row, column=col_idx + 2, value=val)
                if col_idx >= 1:
                    cell.number_format = CURRENCY_FMT
                emi_totals[col_idx] += val
            row += 1

        # Total emitidas
        ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
        for col_idx, val in enumerate(emi_totals):
            cell = ws.cell(row=row, column=col_idx + 2, value=val)
            cell.font = TOTAL_FONT
            if col_idx >= 1:
                cell.number_format = CURRENCY_FMT
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = Border(top=Side(style="thin", color="2F5496"))
        row += 2

        # === SECCIÓN RECIBIDAS ===
        cell = ws.cell(row=row, column=1, value="RECIBIDAS")
        cell.font = SECTION_FONT
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = SECTION_FILL
        row += 1

        for col_idx, (h, _) in enumerate(zip(emi_headers, emi_widths), 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row += 1

        rec_totals = [0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        for month in range(1, 13):
            sr = self.db.monthly_summary(year, month, "recibida")
            ws.cell(row=row, column=1, value=MESES[month])
            values = [
                sr["num_cfdis"], sr["ingresos"], sr["egresos"],
                sr["iva_trasladado"], sr["isr_retenido"], sr["iva_retenido"],
                sr["total"],
            ]
            for col_idx, val in enumerate(values):
                cell = ws.cell(row=row, column=col_idx + 2, value=val)
                if col_idx >= 1:
                    cell.number_format = CURRENCY_FMT
                rec_totals[col_idx] += val
            row += 1

        # Total recibidas
        ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
        for col_idx, val in enumerate(rec_totals):
            cell = ws.cell(row=row, column=col_idx + 2, value=val)
            cell.font = TOTAL_FONT
            if col_idx >= 1:
                cell.number_format = CURRENCY_FMT
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = Border(top=Side(style="thin", color="2F5496"))
        row += 2

        # === SECCIÓN IMPUESTOS PROVISIONALES ===
        fiscal = calcular_impuestos_mensuales(self.db, year)

        FISCAL_FILL = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        FISCAL_SECTION_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        cell = ws.cell(row=row, column=1, value="IMPUESTOS PROVISIONALES (estimado)")
        cell.font = Font(bold=True, size=12, color="C65911")
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = FISCAL_SECTION_FILL
        row += 1

        fisc_headers = [
            "Mes", "IVA Cobrado", "IVA Acredit.", "IVA Retenido",
            "IVA x Pagar", "Base ISR Acum.", "ISR Prov.", "Total x Pagar",
        ]
        fisc_widths = [14, 16, 16, 14, 16, 16, 16, 16]
        for col_idx, (h, w) in enumerate(zip(fisc_headers, fisc_widths), 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = FISCAL_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            if ws.column_dimensions[get_column_letter(col_idx)].width < w:
                ws.column_dimensions[get_column_letter(col_idx)].width = w
        row += 1

        total_iva_pagar = 0.0
        total_isr_prov = 0.0
        for fi in fiscal:
            ws.cell(row=row, column=1, value=MESES[fi["mes"]])
            total_pagar = fi["iva_a_pagar"] + fi["isr_provisional"]
            values = [
                fi["iva_cobrado"], fi["iva_acreditable"], fi["iva_retenido"],
                fi["iva_a_pagar"], fi["base_gravable"], fi["isr_provisional"],
                total_pagar,
            ]
            for col_idx, val in enumerate(values):
                cell = ws.cell(row=row, column=col_idx + 2, value=val)
                cell.number_format = CURRENCY_FMT
            total_iva_pagar += fi["iva_a_pagar"]
            total_isr_prov += fi["isr_provisional"]
            row += 1

        # Total impuestos
        ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
        for col_idx, val in [
            (4, total_iva_pagar), (6, total_isr_prov),
            (7, total_iva_pagar + total_isr_prov),
        ]:
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = TOTAL_FONT
            cell.number_format = CURRENCY_FMT
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = Border(top=Side(style="thin", color="C65911"))
        row += 1

        # Nota
        ws.cell(row=row + 1, column=1, value=(
            "* IVA x Pagar = IVA cobrado - IVA acreditable - IVA retenido"
        )).font = Font(italic=True, size=9, color="888888")
        ws.cell(row=row + 2, column=1, value=(
            "* ISR Prov. = Art.96 LISR sobre (ingresos - deducciones) - ISR retenido - pagos prov. anteriores (no incluye depreciaciones, PTU ni pérdidas ant.)"
        )).font = Font(italic=True, size=9, color="888888")

        ws.freeze_panes = "A2"


    def _get_gastos_deducibles(self, fecha_inicio: date, fecha_fin: date):
        """Obtiene facturas recibidas vigentes tipo I/E (gastos) en un rango."""
        # Solo tipo I (Ingreso=gasto del receptor) y E (Egreso=nota de crédito)
        # Excluye N (Nómina=ingreso), P (Pago=complemento), T (Traslado)
        gastos = []
        for tipo_comp in ("I", "E"):
            gastos.extend(self.db.search(
                tipo="recibida",
                tipo_comprobante=tipo_comp,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                estado="Vigente",
                limit=10000,
            ))
        return gastos

    def _write_fiscal_analysis(self, ws, fecha_inicio: date, fecha_fin: date, titulo: str):
        """Escribe la pestaña de análisis fiscal con clasificación de deducciones."""
        clasificador = ClasificadorDeducciones(self.regimen, self.db)
        recibidas = self._get_gastos_deducibles(fecha_inicio, fecha_fin)

        # Título
        title_cell = ws.cell(row=1, column=1, value=f"Análisis Fiscal - {titulo}")
        title_cell.font = Font(bold=True, size=16, color="283593")

        if not recibidas:
            ws.cell(row=3, column=1, value="Sin facturas recibidas para analizar.").font = Font(
                italic=True, color="888888"
            )
            return

        # ── Sección 1: Resumen por Categoría ──
        row = 3
        cell = ws.cell(row=row, column=1, value="RESUMEN POR CATEGORÍA DE GASTO")
        cell.font = SUGERENCIA_FONT
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = SUGERENCIA_FILL
        row += 1

        cat_headers = [
            ("Categoría", 32), ("Conceptos", 12), ("Monto Total", 16),
            ("Monto Deducible", 16), ("No Deducible", 16), ("% Deducible", 12),
            ("Alertas", 6),
        ]
        for col_idx, (h, w) in enumerate(cat_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        row += 1

        resumen = clasificador.resumen_periodo(recibidas)

        cats_sorted = sorted(
            resumen["por_categoria"].items(),
            key=lambda x: float(x[1]["monto_deducible"]),
            reverse=True,
        )

        for cat_id, cat_data in cats_sorted:
            monto_orig = float(cat_data["monto_original"])
            monto_ded = float(cat_data["monto_deducible"])
            monto_no_ded = monto_orig - monto_ded
            pct = cat_data["porcentaje"]
            n_alertas = len(cat_data["alertas"])

            ws.cell(row=row, column=1, value=cat_data["nombre"])
            ws.cell(row=row, column=2, value=cat_data["num_conceptos"]).alignment = Alignment(horizontal="center")

            for col, val in [(3, monto_orig), (4, monto_ded), (5, monto_no_ded)]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.number_format = CURRENCY_FMT

            pct_cell = ws.cell(row=row, column=6, value=pct / 100)
            pct_cell.number_format = '0%'
            pct_cell.alignment = Alignment(horizontal="center")

            if n_alertas:
                ws.cell(row=row, column=7, value=n_alertas).alignment = Alignment(horizontal="center")

            # Colorear fila según deducibilidad
            if pct >= 100:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = DEDUCIBLE_FILL
            elif pct == 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = NO_DEDUCIBLE_FILL
            elif n_alertas > 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = ALERTA_FILL

            row += 1

        # Fila de totales
        total_orig = float(resumen["total_original"])
        total_ded = float(resumen["total_deducible"])
        total_no_ded = float(resumen["total_no_deducible"])
        pct_global = resumen["porcentaje_global"]

        ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
        for col, val in [(3, total_orig), (4, total_ded), (5, total_no_ded)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = CURRENCY_FMT
            cell.font = TOTAL_FONT
        pct_cell = ws.cell(row=row, column=6, value=pct_global / 100)
        pct_cell.number_format = '0%'
        pct_cell.font = TOTAL_FONT
        pct_cell.alignment = Alignment(horizontal="center")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = Border(top=Side(style="thin", color="283593"))
        row += 2

        # ── Sección 2: Detalle por Concepto ──
        cell = ws.cell(row=row, column=1, value="DETALLE POR CONCEPTO")
        cell.font = SUGERENCIA_FONT
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = SUGERENCIA_FILL
        row += 1

        det_headers = [
            ("Fecha", 12), ("Emisor", 25), ("Concepto", 35), ("Clave SAT", 12),
            ("Categoría", 22), ("Monto", 14), ("Deducible", 14),
            ("%", 8), ("Fundamento", 28),
        ]
        for col_idx, (h, w) in enumerate(det_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            if col_idx > 7:
                ws.column_dimensions[get_column_letter(col_idx)].width = w
        row += 1

        count = 0
        for comp in recibidas:
            if not comp.conceptos:
                continue
            clasificaciones = clasificador.clasificar_comprobante(comp)
            for clas in clasificaciones:
                ws.cell(row=row, column=1, value=comp.fecha.strftime("%d/%m/%Y"))
                ws.cell(row=row, column=2, value=(comp.nombre_emisor or comp.rfc_emisor)[:25])
                ws.cell(row=row, column=3, value=clas.concepto_descripcion[:35])
                ws.cell(row=row, column=4, value=clas.clave_prod_serv)
                ws.cell(row=row, column=5, value=clas.categoria[:22])

                for col, val in [(6, float(clas.monto_original)), (7, float(clas.monto_deducible))]:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.number_format = CURRENCY_FMT

                pct_cell = ws.cell(row=row, column=8, value=clas.porcentaje_deducible / 100)
                pct_cell.number_format = '0%'
                pct_cell.alignment = Alignment(horizontal="center")

                ws.cell(row=row, column=9, value=clas.fundamento_legal)

                # Colorear según resultado
                if not clas.es_deducible:
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = NO_DEDUCIBLE_FILL
                elif clas.alertas:
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = ALERTA_FILL

                row += 1
                count += 1

        # Nota al pie
        row += 1
        ws.cell(row=row, column=1, value=(
            "Verde = 100% deducible | Amarillo = deducible con alertas | "
            "Rosa = no deducible"
        )).font = Font(italic=True, size=9, color="888888")
        row += 1
        ws.cell(row=row, column=1, value=(
            "* Este análisis es educativo y no sustituye asesoría fiscal profesional."
        )).font = Font(italic=True, size=9, color="888888")

        ws.freeze_panes = "A2"

    def _write_suggestions(self, ws, fecha_inicio: date, fecha_fin: date, titulo: str):
        """Escribe la pestaña de sugerencias de optimización fiscal."""
        clasificador = ClasificadorDeducciones(self.regimen, self.db)
        recibidas = self._get_gastos_deducibles(fecha_inicio, fecha_fin)

        # Título
        title_cell = ws.cell(row=1, column=1, value=f"Sugerencias de Optimización Fiscal - {titulo}")
        title_cell.font = Font(bold=True, size=16, color="283593")

        if not recibidas:
            ws.cell(row=3, column=1, value="Sin facturas recibidas para analizar.").font = Font(
                italic=True, color="888888"
            )
            return

        # Calcular ingresos del periodo
        ingresos = 0.0
        emitidas = self.db.search(
            tipo="emitida",
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estado="Vigente",
            limit=10000,
        )
        for e in emitidas:
            if e.tipo_comprobante == "I":
                ingresos += float(e.total)

        sugerencias = clasificador.generar_sugerencias(recibidas, ingresos)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 14

        # ── Resumen rápido ──
        row = 3
        resumen = clasificador.resumen_periodo(recibidas)

        cell = ws.cell(row=row, column=1, value="RESUMEN FISCAL")
        cell.font = SUGERENCIA_FONT
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = SUGERENCIA_FILL
        row += 1

        stats = [
            ("Ingresos del periodo (emitidas)", ingresos),
            ("Total gastos (recibidas)", float(resumen["total_original"])),
            ("Total deducible", float(resumen["total_deducible"])),
            ("Total NO deducible", float(resumen["total_no_deducible"])),
        ]
        for label, val in stats:
            ws.cell(row=row, column=2, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=3, value=val)
            cell.number_format = CURRENCY_FMT
            row += 1

        pct = resumen["porcentaje_global"]
        ws.cell(row=row, column=2, value="Porcentaje deducible").font = Font(bold=True)
        pct_cell = ws.cell(row=row, column=3, value=pct / 100)
        pct_cell.number_format = '0.0%'
        row += 1

        if ingresos > 0:
            tasa_efectiva = float(resumen["total_deducible"]) / ingresos
            ws.cell(row=row, column=2, value="Deducciones / Ingresos").font = Font(bold=True)
            cell = ws.cell(row=row, column=3, value=tasa_efectiva)
            cell.number_format = '0.0%'
            row += 1

        row += 1

        # ── Tabla de sugerencias ──
        cell = ws.cell(row=row, column=1, value="SUGERENCIAS DE OPTIMIZACIÓN")
        cell.font = SUGERENCIA_FONT
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = SUGERENCIA_FILL
        row += 1

        sug_headers = [("#", 6), ("Título", 40), ("Descripción", 80), ("Ahorro Estimado", 18), ("Prioridad", 14)]
        for col_idx, (h, _) in enumerate(sug_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        row += 1

        if not sugerencias:
            ws.cell(row=row, column=2, value="No se encontraron sugerencias adicionales.").font = Font(
                italic=True, color="888888"
            )
            row += 1
        else:
            for i, sug in enumerate(sugerencias, 1):
                ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=2, value=sug.titulo).font = Font(bold=True)

                # Descripción con wrap
                desc_cell = ws.cell(row=row, column=3, value=sug.descripcion)
                desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

                if sug.ahorro_estimado:
                    cell = ws.cell(row=row, column=4, value=float(sug.ahorro_estimado))
                    cell.number_format = CURRENCY_FMT
                    cell.font = FISCAL_GREEN

                prioridad_labels = {1: "Alta", 2: "Media", 3: "Baja"}
                prio_cell = ws.cell(row=row, column=5, value=prioridad_labels.get(sug.prioridad, ""))
                prio_cell.alignment = Alignment(horizontal="center")
                if sug.prioridad == 1:
                    prio_cell.font = FISCAL_RED
                elif sug.prioridad == 2:
                    prio_cell.font = FISCAL_ORANGE

                # Colorear fila por prioridad
                if sug.prioridad == 1:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = NO_DEDUCIBLE_FILL
                elif sug.prioridad == 2:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = ALERTA_FILL

                row += 1

        # ── Alertas del análisis ──
        if resumen["alertas"]:
            row += 1
            cell = ws.cell(row=row, column=1, value="ALERTAS DETECTADAS")
            cell.font = Font(bold=True, size=12, color="C62828")
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = NO_DEDUCIBLE_FILL
            row += 1

            for alerta in resumen["alertas"]:
                ws.cell(row=row, column=2, value=alerta)
                row += 1

        # Nota al pie
        row += 2
        ws.cell(row=row, column=1, value=(
            "* Estas sugerencias son educativas y no sustituyen asesoría fiscal profesional."
        )).font = Font(italic=True, size=9, color="888888")
        row += 1
        ws.cell(row=row, column=1, value=(
            "* Los ahorros estimados se calculan asumiendo una tasa marginal de ISR del 30%."
        )).font = Font(italic=True, size=9, color="888888")

        ws.freeze_panes = "A2"


def _next_month(year: int, month: int) -> date:
    if month == 12:
        return date(year + 1, 1, 1)
    return date(year, month + 1, 1)
